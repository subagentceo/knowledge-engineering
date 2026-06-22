# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl", "pillow"]
# ///
"""
excel_exporter.py — write a Carta-branded .xlsx from report_processor.py output.

Usage (pipe from report_processor.py):
    uv run report_processor.py <<'EOF' | \\
        uv run excel_exporter.py \\
            --title "Securities Ledger Report" \\
            --as-of-date 2024-01-15 \\
            --generated-by "Jane Doe" \\
            --output /tmp/securities_ledger.xlsx
    {"download_url": "...", "columns": [...]}
    EOF

Reads processed report JSON from stdin (output of report_processor.py):
    {
      "data":  {sheet_name: {"columns": [...], "rows": [...]}},
      "stats": {sheet_name: {...}}
    }

Prints the output file path on success.
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_BG  = "C6EBF4"
HEADER_FG  = "2F3943"
SUMMARY_BG = "E8F5E9"
LOGO_PATH  = Path(__file__).parent.parent / "assets" / "Carta_Logo.png"

FORMAT_MAP = {
    "money":      "$#,##0.00",  # overridden per-cell when currency differs
    "percentage": "0.00%",
    "integer":    "#,##0",
    "date":       "mmm d, yyyy",
    "decimal":    "#,##0.0000",
}

# Excel forbids [ ] \ / ? * : in sheet tab names.
_EXCEL_INVALID = str.maketrans("", "", "[]\\/?*:")


def _sanitize_sheet_name(name: str) -> str:
    sanitized = name.translate(_EXCEL_INVALID).strip()
    return (sanitized or "Sheet")[:31]

# Date columns arrive from api-reporting-service JSON as ISO-8601 strings — every
# date-styled column serializes via LocalDate.toString() (audited under CAR-7203).
# The exporter is the presentation layer, so it converts them to real datetimes —
# see _parse_date.
#
# The ambiguous "%m/%d/%Y" is deliberately excluded: nothing upstream emits ambiguous
# slash dates, and accepting it would silently read "06/07/2026" as MM/DD when it might
# be DD/MM — corrupting a customer-facing date. Unlike report_processor._parse_date
# (which also parses user-supplied filter thresholds and so must stay permissive), this
# only ever sees API cell values, so it can be strict. The remaining non-ISO formats are
# unambiguous (year-first or spelled-month) and kept purely as defensive fallbacks.
_DATE_INPUT_FORMATS = ("%Y-%m-%d", "%Y/%m/%d", "%b %d, %Y", "%B %d, %Y")


def _parse_date(val):
    """Parse an interchange date string to a datetime so the cell's `mmm d, yyyy`
    number-format actually renders (Excel no-ops a number-format on a text cell).
    Returns None if unrecognized — the caller then leaves the original string
    untouched, so non-date values (e.g. "N/A") are never lost. Independently safe
    on None / non-str input."""
    if val is None:
        return None
    text = str(val).strip()
    for fmt in _DATE_INPUT_FORMATS:
        try:
            return datetime.strptime(text, fmt)
        except (ValueError, TypeError):
            continue
    return None


def _money_format(symbol: str) -> str:
    """Return an Excel number format string for the given currency symbol."""
    if symbol == "$":
        return "$#,##0.00"
    return f'"{symbol}"#,##0.00'
NUMBER_TYPES = {"money", "integer", "percentage", "decimal"}

EXCEL_AGG_FN = {
    "sum":   "SUM",
    "avg":   "AVERAGE",
    "min":   "MIN",
    "max":   "MAX",
    "count": "COUNTA",
}


def _summary_formula(op, col_letter, start_row, end_row):
    fn = EXCEL_AGG_FN.get(op)
    if not fn or start_row > end_row:
        return None
    return f"={fn}({col_letter}{start_row}:{col_letter}{end_row})"


def _write_sheet(ws, columns, rows, title, as_of_date, generated_by, summary_meta=None,
                 currency=None, row_currencies=None):
    base_symbol    = (currency or {}).get("symbol", "$")
    base_money_fmt = _money_format(base_symbol)
    row_currencies = row_currencies or []

    last_col = get_column_letter(len(columns))

    # Hide gridlines — must use sheetView[0], not ws.sheet_view
    ws.views.sheetView[0].showGridLines = False

    # Row 1: thin top margin
    ws.row_dimensions[1].height = 6

    # Row 2: logo + title
    ws.row_dimensions[2].height = 53  # TITLE_ROW_HEIGHT_IN_POINTS
    title_cell = ws.cell(row=2, column=2, value=title)
    title_cell.font = Font(name="Arial", bold=True, size=16, color=HEADER_FG)
    title_cell.alignment = Alignment(vertical="center")

    try:
        img = XLImage(str(LOGO_PATH))
        img.width = 120
        img.height = 50
        img.anchor = "A2"
        ws.add_image(img)
    except FileNotFoundError:
        pass

    # Row 3: subtitle
    ws.row_dimensions[3].height = 21  # AS_OF_DATE_ROW_HEIGHT_IN_POINTS
    try:
        d = datetime.strptime(as_of_date, "%Y-%m-%d")
        as_of_display = f"{d.strftime('%b')} {d.day}, {d.year}"
    except ValueError:
        raise ValueError(f"as_of_date must be in YYYY-MM-DD format, got: {as_of_date!r}")
    n = datetime.now().astimezone()
    now = f"{n.strftime('%b')} {n.day}, {n.year} {n.strftime('%I:%M:%S %p %Z')}"
    sub = ws.cell(row=3, column=2,
                  value=f"As of {as_of_display} • Generated with Claude AI by {generated_by} at {now} • Date format: MMM D, YYYY")
    sub.font = Font(name="Arial", size=10, color="666666")

    # Row 4: blank separator
    ws.row_dimensions[4].height = 15

    # Row 5: headers + auto-filter
    header_fill = PatternFill("solid", fgColor=HEADER_BG)
    header_font = Font(name="Arial", bold=True, size=12, color=HEADER_FG)
    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=5, column=col_idx, value=col["name"] if not col.get("_synthetic") else "")
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = (
            18 if col["type"] in NUMBER_TYPES else 35
        )
    ws.auto_filter.ref = f"A5:{last_col}5"

    # Row 6+: data
    DATA_START = 6
    summary_count = summary_meta["count"] if summary_meta else 0
    data_row_count = len(rows) - summary_count
    last_data_row  = DATA_START + data_row_count - 1

    summary_fill = PatternFill("solid", fgColor=SUMMARY_BG)
    summary_font = Font(name="Arial", bold=True, size=11)
    agg_ops      = summary_meta.get("ops", {}) if summary_meta else {}

    section_fill = PatternFill("solid", fgColor="F7F8F8")
    section_font = Font(name="Arial", bold=True, size=10, color="656B6B")

    for row_idx, row in enumerate(rows, start=DATA_START):
        row_data_idx = row_idx - DATA_START  # 0-based position in rows list

        if isinstance(row, dict):
            if "_s" in row:
                cell = ws.cell(row=row_idx, column=1, value=row["_s"])
                cell.font = section_font
                cell.fill = section_fill
                continue
            elif "_t" in row and isinstance(row["_t"], list):
                row = row["_t"]
            else:
                continue
        elif isinstance(row, list) and len(row) == 1 and isinstance(row[0], str):
            # Plain single-element string arrays are section headers from the API
            cell = ws.cell(row=row_idx, column=1, value=row[0])
            cell.font = section_font
            cell.fill = section_fill
            continue

        is_summary = summary_count > 0 and (row_idx - DATA_START) >= data_row_count

        # row_curr can be: None | {code,symbol} (row-level) | [None|{code,symbol},...] (cell-level)
        row_curr = row_currencies[row_data_idx] if row_data_idx < len(row_currencies) else None

        for col_idx, (col, val) in enumerate(zip(columns, row), start=1):
            if is_summary:
                col_letter = get_column_letter(col_idx)
                op = agg_ops.get(col["name"])
                if col_idx == 1:
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                elif op:
                    formula = _summary_formula(op, col_letter, DATA_START, last_data_row)
                    cell = ws.cell(row=row_idx, column=col_idx, value=formula)
                else:
                    cell = ws.cell(row=row_idx, column=col_idx, value=None)
                cell.fill = summary_fill
                cell.font = summary_font
            else:
                if col["type"] == "integer" and val is not None:
                    try:
                        val = int(val)
                    except (ValueError, TypeError):
                        pass
                elif col["type"] == "money" and val is not None:
                    try:
                        val = float(val)
                    except (ValueError, TypeError):
                        pass
                elif col["type"] in ("percentage", "decimal") and val is not None:
                    try:
                        # API may return percentages as "46.81%" strings or decimal strings
                        if isinstance(val, str) and val.endswith('%'):
                            val = float(val.rstrip('%')) / 100
                        else:
                            val = float(val)
                    except (ValueError, TypeError):
                        pass
                elif col["type"] == "date" and isinstance(val, str) and val.strip():
                    # Coerce interchange date strings to real datetimes so FORMAT_MAP["date"]
                    # ("mmm d, yyyy") renders. Unparseable strings fall through unchanged.
                    parsed = _parse_date(val)
                    if parsed is not None:
                        val = parsed
                cell = ws.cell(row=row_idx, column=col_idx, value=val)

            if col["type"] == "money":
                if is_summary:
                    cell.number_format = base_money_fmt
                else:
                    zero_idx = col_idx - 1
                    if isinstance(row_curr, list):
                        cell_curr = row_curr[zero_idx] if zero_idx < len(row_curr) else None
                        sym = cell_curr["symbol"] if cell_curr else base_symbol
                    elif row_curr:
                        sym = row_curr["symbol"]
                    else:
                        sym = base_symbol
                    cell.number_format = _money_format(sym)
            elif col["type"] in FORMAT_MAP:
                cell.number_format = FORMAT_MAP[col["type"]]

    ws.freeze_panes = "A6"
    ws.sheet_view.selection[0].activeCell = "A6"
    ws.sheet_view.selection[0].sqref = "A6"


def main():
    parser = argparse.ArgumentParser(description="Export report JSON to Carta-branded .xlsx")
    parser.add_argument("--title",        required=True,  help="Report title shown in cell B2")
    parser.add_argument("--as-of-date",   required=True,  help="As-of date in YYYY-MM-DD format (output renders as MMM D, YYYY per Carta standard)")
    parser.add_argument("--generated-by", required=True,  help="User name for the subtitle line")
    parser.add_argument("--output",       required=True,  help="Output .xlsx file path")
    args = parser.parse_args()

    payload = json.load(sys.stdin)
    data    = payload.get("data", payload)  # accept both full payload and bare data dict

    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    for sheet_name, sheet in data.items():
        ws = wb.create_sheet(title=_sanitize_sheet_name(sheet_name))
        _write_sheet(
            ws,
            columns=sheet["columns"],
            rows=sheet["rows"],
            title=args.title,
            as_of_date=args.as_of_date,
            generated_by=args.generated_by,
            summary_meta=sheet.get("summary_meta"),
            currency=sheet.get("currency"),
            row_currencies=sheet.get("row_currencies"),
        )

    out_path = Path(args.output)
    wb.save(out_path)
    print(out_path)


if __name__ == "__main__":
    main()
